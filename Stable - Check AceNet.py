from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import tkinter as tk
from tkinter import messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import Font
import time
import os
import re
from datetime import datetime
import tkinter.simpledialog as simpledialog
import requests
from selenium.common.exceptions import WebDriverException

def prompt_user_login():
    """Prompt user to manually log in to the website."""
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("Login Required", "Please manually log in to the website, then click OK.")
    root.destroy()

def prompt_start_search():
    """Prompt user to confirm they are ready to start the search."""
    root = tk.Tk()
    root.withdraw()
    response = messagebox.askokcancel("Start Search", "Are you ready to start the SKU search?")
    root.destroy()
    return response

def check_file_access(file_path):
    """Check if the file is accessible for reading and writing."""
    try:
        with open(file_path, 'rb') as f:
            pass
        with open(file_path, 'ab') as f:
            pass
        return True
    except Exception as e:
        print(f"File access check failed: {e}")
        return False

def get_part_numbers(excel_file, sheet_name):
    """Read all PARTNUMBERs from the specified macro-enabled Excel sheet."""
    print(f"Reading macro-enabled Excel file: {excel_file}, Sheet: {sheet_name}")
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        if "PARTNUMBER" not in df.columns:
            raise ValueError("Column 'PARTNUMBER' not found in the sheet.")
        def clean_part_number(val):
            if pd.isna(val):
                return None
            if isinstance(val, float) and val.is_integer():
                return str(int(val))
            return str(val)
        part_numbers = [clean_part_number(val) for val in df["PARTNUMBER"].dropna()]
        print(f"Retrieved {len(part_numbers)} PARTNUMBERs")
        return part_numbers
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []

def initialize_excel_output(excel_file, sheet_name):
    """Initialize the output sheet with headers and set initial column widths."""
    print(f"Initializing new Excel file: {excel_file}, Sheet: {sheet_name}")
    for attempt in range(5):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Initialize headers in columns A, C, E, G, I, K, M (blank columns B, D, F, H, J, L)
            ws['A1'] = "No Discovery"
            ws['C1'] = "No Asterisk(*)"
            ws['E1'] = "Cancelled"
            ws['G1'] = "On Order"
            ws['I1'] = "No Location"
            ws['K1'] = "Not in AceNet"
            ws['M1'] = "Not in RSC"
            for col in ['A', 'C', 'E', 'G', 'I', 'K', 'M']:
                ws[f'{col}1'].font = Font(bold=True)
                ws.column_dimensions[col].width = len(ws[f'{col}1'].value) * 1.2
            # Ensure blank columns have default width
            for col in ['B', 'D', 'F', 'H', 'J', 'L']:
                ws.column_dimensions[col].width = 2  # Minimal width for blank columns
            
            wb.save(excel_file)
            print(f"Initialized '{sheet_name}' with headers in {excel_file}")
            return
        except Exception as e:
            print(f"Attempt {attempt + 1}/5: Error initializing Excel output: {e}")
            time.sleep(2)
    print("Failed to initialize Excel output after 5 attempts.")
    print(f"Please ensure you have write permissions for the Desktop and no application is locking '{excel_file}'.")
    raise Exception("Cannot initialize Excel output due to file access issues.")

def append_to_excel(excel_file, sheet_name, column, part_number):
    """Append a PARTNUMBER to the specified column and auto-size the column."""
    column_map = {
        "No Discovery": 'A',
        "No Asterisk(*)": 'C',
        "Cancelled": 'E',
        "On Order": 'G',
        "No Location": 'I',
        "Not in AceNet": 'K',
        "Not in RSC": 'M'
    }
    col_letter = column_map.get(column)
    if not col_letter:
        print(f"Error: Invalid column '{column}'")
        return
    
    print(f"Appending to Excel file: {excel_file}, Sheet: {sheet_name}, Column: {col_letter}")
    for attempt in range(5):
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb[sheet_name]
            
            # Find the first empty row in the column (starting from row 2)
            current_row = 2
            while ws[f'{col_letter}{current_row}'].value:
                current_row += 1
            ws[f'{col_letter}{current_row}'] = part_number
            
            # Auto-size column based on max content length
            header = ws[f'{col_letter}1'].value
            max_length = max(len(str(header)), len(str(part_number)))
            for row in range(2, current_row + 1):
                cell_value = ws[f'{col_letter}{row}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = max_length * 1.2
            
            wb.save(excel_file)
            print(f"Appended PARTNUMBER '{part_number}' to column {col_letter}, row {current_row}, column width set to {max_length * 1.2}")
            return current_row
        except Exception as e:
            print(f"Attempt {attempt + 1}/5: Error appending to Excel: {e}")
            time.sleep(2)
    print(f"Failed to append PARTNUMBER '{part_number}' to column '{column}' after 5 attempts.")
    print(f"Please ensure you have write permissions for the Desktop and no application is locking '{excel_file}'.")

def color_code_multi_column_entries(excel_file, sheet_name):
    """Color-code PARTNUMBERs that appear in multiple columns with red text."""
    print(f"Color-coding multi-column PARTNUMBERs in {excel_file}, Sheet: {sheet_name}")
    for attempt in range(5):
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb[sheet_name]
            
            # Collect PARTNUMBERs and their columns (A, C, E, G, I)
            part_number_columns = {}
            for col in ['A', 'C', 'E', 'G', 'I']:
                row = 2
                while ws[f'{col}{row}'].value:
                    part_number = ws[f'{col}{row}'].value
                    if part_number not in part_number_columns:
                        part_number_columns[part_number] = []
                    part_number_columns[part_number].append((col, row))
                    row += 1
            
            # Apply red color to PARTNUMBERs in multiple columns
            red_font = Font(color="FF0000")
            for part_number, locations in part_number_columns.items():
                if len(locations) > 1:  # PARTNUMBER appears in multiple columns
                    for col, row in locations:
                        ws[f'{col}{row}'].font = red_font
                        print(f"Colored PARTNUMBER '{part_number}' red in column {col}, row {row}")
            
            wb.save(excel_file)
            print("Color-coding completed successfully.")
            return
        except Exception as e:
            print(f"Attempt {attempt + 1}/5: Error color-coding Excel: {e}")
            time.sleep(2)
    print("Failed to color-code Excel after 5 attempts.")
    print(f"Please ensure you have write permissions for the Desktop and no application is locking '{excel_file}'.")

def create_status_popup(total_items):
    """Create a non-blocking Tkinter status popup."""
    root = tk.Tk()
    root.title("Processing PARTNUMBERs")
    root.geometry("300x100")
    status_label = tk.Label(root, text="Initializing...", font=("Arial", 12))
    status_label.pack(pady=20)
    
    def update_status(current, total):
        status_label.config(text=f"{current} of {total} items checked")
        root.update()
    
    def close_popup():
        root.destroy()
    
    return update_status, close_popup

def create_completion_popup(excel_file):
    """Create a popup after completion with options to close or open the Excel file."""
    root = tk.Tk()
    root.title("Processing Complete")
    root.geometry("300x150")
    
    label = tk.Label(root, text="All items checked", font=("Arial", 12))
    label.pack(pady=20)
    
    button_frame = tk.Frame(root)
    button_frame.pack(pady=10)
    
    def close_popup():
        root.destroy()
    
    def open_excel():
        try:
            os.startfile(excel_file)  # Windows-specific to open file in default application
            print(f"Opened Excel file: {excel_file}")
        except Exception as e:
            print(f"Error opening Excel file: {e}")
            messagebox.showerror("Error", f"Failed to open Excel file: {e}")
        root.destroy()
    
    close_button = tk.Button(button_frame, text="Close", command=close_popup, width=10)
    close_button.pack(side=tk.LEFT, padx=5)
    
    open_button = tk.Button(button_frame, text="Open Excel", command=open_excel, width=10)
    open_button.pack(side=tk.LEFT, padx=5)
    
    root.mainloop()

def prompt_for_credentials():
    root = tk.Tk()
    root.withdraw()
    username = simpledialog.askstring("Login", "Enter your AceNet username:")
    password = simpledialog.askstring("Login", "Enter your AceNet password:", show='*')
    # Store selection dropdown
    store_numbers = ["16719", "17521", "18179", "18181"]
    store_selection = tk.StringVar(value=store_numbers[0])
    def select_store():
        store_window.destroy()
    store_window = tk.Toplevel(root)
    store_window.title("Select Store")
    tk.Label(store_window, text="Select your store number:").pack(pady=10)
    for sn in store_numbers:
        tk.Radiobutton(store_window, text=sn, variable=store_selection, value=sn).pack(anchor=tk.W)
    tk.Button(store_window, text="OK", command=select_store).pack(pady=10)
    store_window.grab_set()
    root.wait_window(store_window)
    store = store_selection.get()
    root.destroy()
    return username, password, store

def login_and_select_store(driver, username, password, store):
    driver.get("https://acenet.aceservices.com/")
    try:
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "userNameInput"))).send_keys(username)
        password_elem = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "passwordInput")))
        password_elem.send_keys(password)
        password_elem.send_keys(Keys.RETURN)
        # Wait for store dropdown to be clickable
        store_dropdown = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "filter-option-inner-inner"))
        )
        store_dropdown.click()
        # Wait for the dropdown list to appear
        dropdown_ul = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div[role='listbox'] ul.dropdown-menu"))
        )
        # Find the <a> element with the correct store number
        store_a = None
        store_items = dropdown_ul.find_elements(By.TAG_NAME, "a")
        for a in store_items:
            try:
                span = a.find_element(By.CSS_SELECTOR, "span.store-number")
                if span.text.strip().startswith(store):
                    store_a = a
                    break
            except Exception:
                continue
        if store_a:
            store_a.click()
        else:
            print(f"Store number {store} not found in dropdown!")
            raise Exception(f"Store number {store} not found in dropdown!")
        # Wait for the page to refresh and the search box to appear
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "tbxSearchBox")))
    except Exception as e:
        print(f"Login or store selection failed: {e}")
        raise

def is_driver_alive(driver):
    try:
        _ = driver.title
        return True
    except WebDriverException:
        return False

def automate_search():
    """Automate SKU search to find PARTNUMBERs with no content in the Discovery element, no asterisk in the linked element, cancelled status, on order, or with no location text."""
    driver = None
    max_restarts = 50
    restart_count = 0
    try:
        # Prompt for credentials and store selection
        username, password, store = prompt_for_credentials()
        # Resolve desktop path
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        input_excel = os.path.join(desktop_path, "Project FWN.xlsm")
        output_excel = os.path.join(desktop_path, f"No Discovery Check {datetime.now().strftime('%Y-%m-%d')}.xlsx")
        input_sheet = "Big Beautiful Order"
        output_sheet = "No Discovery Check"
        print("Verifying macro-enabled input Excel file exists...")
        if not os.path.exists(input_excel):
            print(f"Error: Macro-enabled Excel file '{input_excel}' not found on Desktop.")
            return
        print("Checking input file access...")
        if not check_file_access(input_excel):
            print("Error: Cannot access 'Project FWN.xlsm'. Ensure it's not open in Excel and you have read permissions.")
            return
        print("Reading PARTNUMBERs...")
        part_numbers = get_part_numbers(input_excel, input_sheet)
        if not part_numbers:
            print("Error: No PARTNUMBERs found. Check sheet and column.")
            return
        print("Initializing new Excel output file...")
        initialize_excel_output(output_excel, output_sheet)
        # Initialize status popup
        update_status, close_popup = create_status_popup(len(part_numbers))
        chrome_options = Options()
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--disable-web-security")
        chrome_options.add_argument("--allow-running-insecure-content")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36")
        chrome_options.add_argument("--disable-features=VoiceTranscription,SpeechRecognition")
        chrome_options.add_argument("--disable-media-stream")
        chrome_options.add_argument("--disable-webrtc")
        service = Service(ChromeDriverManager().install())
        def start_browser_and_login():
            driver = webdriver.Chrome(service=service, options=chrome_options)
            login_and_select_store(driver, username, password, store)
            return driver
        driver = start_browser_and_login()
        main_window = driver.current_window_handle
        popup_window = None
        idx = 0
        while idx < len(part_numbers):
            part_number = part_numbers[idx]
            # Check driver health before each PARTNUMBER
            if not is_driver_alive(driver):
                print("WebDriver is not alive. Restarting browser...")
                try:
                    driver.quit()
                except:
                    pass
                restart_count += 1
                if restart_count > max_restarts:
                    print(f"WebDriver has been restarted more than {max_restarts} times. Aborting script.")
                    break
                driver = start_browser_and_login()
                main_window = driver.current_window_handle
            try:
                print(f"Processing PARTNUMBER {idx+1}/{len(part_numbers)}: '{part_number}'")
                update_status(idx+1, len(part_numbers))
                driver.switch_to.window(main_window)
                try:
                    sku_field = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "tbxSearchBox"))
                    )
                except Exception as e:
                    print(f"Search box not found on first try: {e}. Navigating main window to home page and retrying...")
                    driver.switch_to.window(main_window)
                    driver.get("https://acenet.aceservices.com/")
                    WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.ID, "tbxSearchBox"))
                    )
                    time.sleep(2)
                    try:
                        sku_field = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.ID, "tbxSearchBox"))
                        )
                    except Exception as e2:
                        print(f"Search box still not found after navigation: {e2}. Skipping PARTNUMBER '{part_number}'.")
                        append_to_excel(output_excel, output_sheet, "Not in AceNet", part_number)
                        # Clean up popups
                        for handle in driver.window_handles:
                            if handle != main_window:
                                try:
                                    driver.switch_to.window(handle)
                                    driver.close()
                                except:
                                    pass
                        driver.switch_to.window(main_window)
                        idx += 1
                        continue
                sku_field.clear()
                sku_field.send_keys(part_number)
                sku_field.send_keys(Keys.RETURN)
                print("Entered PARTNUMBER in main window input field")
                wait_time = 10 if idx == 0 else 1
                print(f"Waiting {wait_time} seconds for popup window...")
                time.sleep(wait_time)
                print("--- DEBUG: Window handles, titles, and URLs after search ---")
                try:
                    for handle in driver.window_handles:
                        try:
                            driver.switch_to.window(handle)
                            print(f"Handle: {handle}, Title: {driver.title}, URL: {driver.current_url}")
                        except Exception as e:
                            print(f"Handle: {handle}, Error getting title/URL: {e}")
                except Exception as e:
                    print(f"Connection error or window handle error: {e}")
                    print(f"Closing all Chrome windows and restarting browser for PARTNUMBER '{part_number}'...")
                    append_to_excel(output_excel, output_sheet, "Not in AceNet", part_number)
                    try:
                        if driver:
                            driver.quit()
                    except:
                        pass
                    # Force kill all Chrome and ChromeDriver processes
                    try:
                        print("Killing all chrome.exe and chromedriver.exe processes...")
                        os.system("taskkill /f /im chrome.exe /T")
                        os.system("taskkill /f /im chromedriver.exe /T")
                        time.sleep(5)
                    except Exception as kill_e:
                        print(f"Error killing Chrome processes: {kill_e}")
                    restart_count += 1
                    if restart_count > max_restarts:
                        print(f"WebDriver has been restarted more than {max_restarts} times. Aborting script.")
                        break
                    try:
                        driver = start_browser_and_login()
                        main_window = driver.current_window_handle
                    except Exception as e2:
                        print(f"Failed to restart browser and login: {e2}")
                        break
                    idx += 1
                    continue
                driver.switch_to.window(main_window)
                print("--- END DEBUG ---")
                main_url = driver.current_url
                if "/search/product?q=" in main_url:
                    print(f"PARTNUMBER '{part_number}' not found in AceNet (main window redirected to search page).")
                    append_to_excel(output_excel, output_sheet, "Not in AceNet", part_number)
                    driver.switch_to.window(main_window)
                    driver.get("https://acenet.aceservices.com/")
                    WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.ID, "tbxSearchBox"))
                    )
                    time.sleep(2)
                    for handle in driver.window_handles:
                        if handle != main_window:
                            try:
                                driver.switch_to.window(handle)
                                driver.close()
                            except:
                                pass
                    driver.switch_to.window(main_window)
                    idx += 1
                    continue
                windows = driver.window_handles
                if len(windows) < 2:
                    print(f"Error: Popup window not found for PARTNUMBER '{part_number}'.")
                    append_to_excel(output_excel, output_sheet, "Not in AceNet", part_number)
                    idx += 1
                    continue
                for window in windows:
                    if window != main_window:
                        popup_window = window
                        driver.switch_to.window(popup_window)
                        print(f"Switched to popup window: {popup_window}")
                        break
                print("Checking for iframes in popup...")
                try:
                    iframes = WebDriverWait(driver, 5).until(
                        EC.presence_of_all_elements_located((By.TAG_NAME, "iframe"))
                    )
                except Exception as e:
                    print(f"No iframes found in popup: {e}")
                    append_to_excel(output_excel, output_sheet, "Not in AceNet", part_number)
                    driver.switch_to.window(main_window)
                    idx += 1
                    continue
                print(f"Found {len(iframes)} iframe(s). Attempting to switch...")
                iframe_switched = False
                for i, iframe in enumerate(iframes):
                    try:
                        driver.switch_to.frame(iframe)
                        print(f"Switched to iframe {i}")
                        try:
                            driver.find_element(By.XPATH, "/html/body/form/div[4]/div[1]/div[11]/div[1]/div[1]/div[20]/div[2]")
                            iframe_switched = True
                            break
                        except Exception as e:
                            print(f"No Discovery element not found in iframe {i}: {e}")
                            print("--- DEBUG: First 500 chars of iframe HTML ---")
                            print(driver.page_source[:500])
                            print("--- END DEBUG ---")
                            driver.switch_to.default_content()
                            continue
                    except Exception as e:
                        print(f"Failed to switch to iframe {i}: {e}")
                        driver.switch_to.default_content()
                if not iframe_switched:
                    print(f"PARTNUMBER '{part_number}' not found in AceNet (no valid iframe content).")
                    append_to_excel(output_excel, output_sheet, "Not in AceNet", part_number)
                    driver.switch_to.window(main_window)
                    idx += 1
                    continue
                is_cancelled = False
                is_not_in_rsc = False
                print("Step 3: Checking Cancelled...")
                try:
                    status_div = WebDriverWait(driver, 10).until(
                        EC.visibility_of_element_located((By.ID, "ctl00_ctl00_contentMainPlaceHolder_MainContent_imagesVideos_mainStatusDiv"))
                    )
                    status_text = status_div.text.strip()
                    print(f"Cancelled text: '{status_text}'")
                    # Check for Not in RSC first
                    if status_text:
                        if re.search(r'not\s+carried\s+in\s+your\s+rsc', status_text, re.IGNORECASE) or \
                           re.search(r'not\s+carried\s+by\s+rsc', status_text, re.IGNORECASE) or \
                           re.search(r'not\s+in\s+rsc', status_text, re.IGNORECASE):
                            print("Outputting PARTNUMBER to Not in RSC due to matching status.")
                            append_to_excel(output_excel, output_sheet, "Not in RSC", part_number)
                            is_not_in_rsc = True
                    # If Not in RSC, skip all other checks for this PARTNUMBER
                    if is_not_in_rsc:
                        if iframe_switched:
                            driver.switch_to.default_content()
                        driver.switch_to.window(main_window)
                        idx += 1
                        continue
                    if status_text:
                        has_cancelled_or_closeout = re.search(r'cancel(?:led|lation)?|close(?:out|-out|d out)', status_text, re.IGNORECASE)
                        has_replacement_or_discontinued = re.search(r'replacement|discontinued', status_text, re.IGNORECASE)
                        if has_cancelled_or_closeout and not has_replacement_or_discontinued:
                            print("Outputting PARTNUMBER to Cancelled due to matching status.")
                            append_to_excel(output_excel, output_sheet, "Cancelled", part_number)
                            is_cancelled = True
                except Exception as e:
                    print(f"Cancelled element not found, skipping: {e}")
                # If Cancelled, skip all other checks for this PARTNUMBER
                if is_cancelled:
                    if iframe_switched:
                        driver.switch_to.default_content()
                    driver.switch_to.window(main_window)
                    idx += 1
                    continue
                print("Step 1: Checking No Discovery element...")
                element_found = False
                discovery_text = ""
                max_attempts = 3 if idx == 0 else 2
                for attempt in range(max_attempts):
                    try:
                        content_div = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, "/html/body/form/div[4]/div[1]/div[11]/div[1]/div[1]/div[20]/div[2]"))
                        )
                        discovery_text = content_div.text.strip()
                        element_found = True
                        print(f"No Discovery element found with text: '{discovery_text}' (attempt {attempt + 1})")
                        break
                    except Exception as e:
                        print(f"No Discovery element not found (attempt {attempt + 1}): {e}")
                        print("--- DEBUG: First 500 chars of iframe HTML (No Discovery step) ---")
                        print(driver.page_source[:500])
                        print("--- END DEBUG ---")
                        if attempt < max_attempts - 1:
                            print("Retrying after 3-second delay...")
                            time.sleep(3)
                        continue
                if (not element_found or not discovery_text.strip()) and not is_cancelled:
                    print("Outputting PARTNUMBER to No Discovery due to element absence or empty text.")
                    append_to_excel(output_excel, output_sheet, "No Discovery", part_number)
                print("Step 2: Checking for non-blank text with no asterisk...")
                try:
                    link_element = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.XPATH, "/html/body/form/div[4]/div[1]/div[11]/div[1]/div[1]/div[20]/div[2]/a"))
                    )
                    link_text = link_element.text.strip()
                    print(f"No Asterisk text: '{link_text}'")
                    if link_text and '*' not in link_text:
                        print("Outputting PARTNUMBER to No Asterisk(*).")
                        append_to_excel(output_excel, output_sheet, "No Asterisk(*)", part_number)
                except:
                    print("No Asterisk element not found, skipping.")
                print("Step 4: Checking On Order...")
                try:
                    order_span = WebDriverWait(driver, 10).until(
                        EC.visibility_of_element_located((By.ID, "spnQOO"))
                    )
                    order_text = order_span.text.strip()
                    print(f"On Order text: '{order_text}'")
                    try:
                        order_value = float(order_text)
                        if order_value > 0:
                            print("Outputting PARTNUMBER to On Order due to value > 0.")
                            append_to_excel(output_excel, output_sheet, "On Order", part_number)
                        else:
                            print("Order value <= 0, skipping.")
                    except ValueError:
                        print("Order text is not a number, skipping.")
                except:
                    print("On Order element not found, skipping.")
                print("Step 5: Checking No Location...")
                try:
                    location_element = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.XPATH, "/html/body/form/div[4]/div[1]/div[11]/div[1]/div[3]/div[17]/div[2]"))
                    )
                    location_text = location_element.text.strip()
                    print(f"No Location text: '{location_text}' (raw: {repr(location_text)})")
                    if not location_text:
                        print(f"Outputting PARTNUMBER to No Location: element text is empty.")
                        append_to_excel(output_excel, output_sheet, "No Location", part_number)
                    else:
                        print(f"Skipping No Location: element contains text '{location_text}'.")
                except Exception as e:
                    print(f"No Location element not found: {str(e)}")
                    print(f"Outputting PARTNUMBER to No Location: element not found.")
                    append_to_excel(output_excel, output_sheet, "No Location", part_number)
                if iframe_switched:
                    driver.switch_to.default_content()
                driver.switch_to.window(main_window)
                idx += 1
            except Exception as e:
                print(f"Error processing PARTNUMBER '{part_number}': {e}")
                append_to_excel(output_excel, output_sheet, "Not in AceNet", part_number)
                try:
                    if driver:
                        driver.quit()
                except:
                    pass
                restart_count += 1
                if restart_count > max_restarts:
                    print(f"WebDriver has been restarted more than {max_restarts} times. Aborting script.")
                    break
                try:
                    driver = start_browser_and_login()
                    main_window = driver.current_window_handle
                except Exception as e2:
                    print(f"Failed to restart browser and login: {e2}")
                    break
                idx += 1
                continue
        print("All PARTNUMBERs processed. Closing status popup...")
        close_popup()
        print("Color-coding multi-column PARTNUMBERs...")
        color_code_multi_column_entries(output_excel, output_sheet)
        print("Showing completion popup...")
        create_completion_popup(output_excel)
        print("Processing completed successfully. Browser will remain open.")
        while True:
            time.sleep(60)
    except Exception as e:
        print(f"Unexpected error occurred: {e}")
        print("Browser will remain open for troubleshooting.")
        if driver:
            while True:
                time.sleep(60)
        else:
            print("Driver not initialized. Script cannot continue.")

if __name__ == "__main__":
    automate_search()